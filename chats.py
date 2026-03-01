import re
import asyncio
from pathlib import Path

from openpyxl import load_workbook
from telethon import TelegramClient
from telethon.errors import FloodWaitError, InviteRequestSentError
from telethon.tl.functions.channels import JoinChannelRequest, LeaveChannelRequest
from telethon.tl.functions.messages import ImportChatInviteRequest, DeleteChatUserRequest
from telethon.tl.types import Channel, Chat, User

from broadcast import create_client
from core import SESSIONS_DIR, load_config, is_account_authorized, add_chat_links, get_chat_links

LINK_PATTERN = re.compile(
    r'(?:https?://)?(?:t\.me|telegram\.me|telegram\.dog)/([a-zA-Z0-9_+/-]+)',
    re.IGNORECASE
)


def extract_links_from_xlsx(filepath: str) -> list[str]:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {filepath}")
    links = set()
    wb = load_workbook(path, data_only=True)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                val = cell.value
                if val is not None:
                    for m in LINK_PATTERN.finditer(str(val)):
                        part = m.group(1).strip("/")
                        links.add(f"t.me/{part}")
                if hasattr(cell, "hyperlink") and cell.hyperlink:
                    href = str(getattr(cell.hyperlink, "target", "") or "")
                    for m in LINK_PATTERN.finditer(href):
                        part = m.group(1).strip("/")
                        links.add(f"t.me/{part}")
    return list(links)


def _parse_link(link: str) -> tuple[str | None, str | None]:
    link = str(link).strip()
    m = re.search(r't\.me/joinchat/([a-zA-Z0-9_-]+)', link, re.I)
    if m:
        return "invite", m.group(1)
    m = re.search(r'telegram\.me/joinchat/([a-zA-Z0-9_-]+)', link, re.I)
    if m:
        return "invite", m.group(1)
    m = re.search(r'(?:t\.me|telegram\.me)/\+([a-zA-Z0-9_-]+)', link, re.I)
    if m:
        return "invite", m.group(1)
    m = re.search(r'(?:t\.me|telegram\.me)/([a-zA-Z0-9_]+)', link, re.I)
    if m:
        return "username", m.group(1)
    return None, None


async def join_by_link(client: TelegramClient, link: str) -> tuple[bool, str]:
    link_type, value = _parse_link(link)
    if not value:
        return False, "Неверный формат ссылки"
    max_retries = 3
    for attempt in range(max_retries):
        try:
            if link_type == "invite":
                await client(ImportChatInviteRequest(value))
            else:
                await client(JoinChannelRequest(value))
            return True, "OK"
        except InviteRequestSentError:
            return True, "Заявка подана (ожидает одобрения)"
        except FloodWaitError as e:
            if attempt < max_retries - 1:
                await asyncio.sleep(e.seconds)
            else:
                return False, f"FloodWait: ждать {e.seconds} сек"
        except Exception as e:
            return False, str(e)
    return False, "Ошибка после повторов"


async def join_links_from_account(
    client: TelegramClient, phone: str, links: list[str], on_progress=None, cancel_event=None
) -> dict:
    joined = 0
    failed = 0
    errors = []
    for link in links:
        if cancel_event and cancel_event.is_set():
            break
        success, msg = await join_by_link(client, link)
        if success:
            joined += 1
        else:
            failed += 1
            errors.append((link, msg))
            try:
                from dashboard import add_alert
                err_l = (msg or "").lower()
                if "auth" in err_l or "session" in err_l or "unregister" in err_l or "не авторизован" in err_l:
                    add_alert("critical", "Аккаунт не авторизован", f"{phone}: {msg}", category="account")
                else:
                    add_alert("warning", "Не удалось вступить в чат", f"{phone} -> {link}: {msg}", category="chat")
            except Exception:
                pass
        if on_progress:
            on_progress(joined, failed)
        await asyncio.sleep(3)
    return {"joined": joined, "failed": failed, "errors": errors}


async def run_join_all_links(links: list[str], on_progress=None, cancel_event=None) -> list[dict]:
    cfg = load_config()
    if not cfg or not cfg.get("api_id") or not cfg.get("api_hash"):
        raise ValueError("Не настроены api_id и api_hash")

    accounts = [a for a in cfg["accounts"] if is_account_authorized(a["phone"])]
    if not accounts:
        raise ValueError("Нет авторизованных аккаунтов")

    account_stats = {}

    async def join_for_account(acc: dict):
        phone = acc["phone"]
        session_path = SESSIONS_DIR / phone.replace("+", "").replace(" ", "")
        proxy = acc.get("proxy") or cfg.get("proxy")
        client = create_client(session_path, cfg["api_id"], cfg["api_hash"], proxy)
        account_stats[phone] = {"joined": 0, "failed": 0}
        try:
            await client.connect()
            if not await client.is_user_authorized():
                return {"phone": phone, "joined": 0, "failed": len(links), "errors": [("", "Не авторизован")]}

            def acc_progress(joined, failed):
                account_stats[phone]["joined"] = joined
                account_stats[phone]["failed"] = failed
                if on_progress:
                    total_j = sum(s["joined"] for s in account_stats.values())
                    total_f = sum(s["failed"] for s in account_stats.values())
                    on_progress(total_j, total_f)

            r = await join_links_from_account(client, phone, links, on_progress=acc_progress, cancel_event=cancel_event)
            return {"phone": phone, **r}
        except Exception as e:
            return {"phone": phone, "joined": 0, "failed": len(links), "errors": [("", str(e))]}
        finally:
            await client.disconnect()

    results = await asyncio.gather(*[join_for_account(acc) for acc in accounts])
    return list(results)


async def _leave_chat(client: TelegramClient, entity) -> tuple[bool, str]:
    try:
        if isinstance(entity, Channel):
            await client(LeaveChannelRequest(entity))
        elif isinstance(entity, Chat):
            me = await client.get_input_entity("me")
            await client(DeleteChatUserRequest(chat_id=entity.id, user_id=me))
        else:
            return False, "Пропуск (не чат/канал)"
        return True, "OK"
    except Exception as e:
        return False, str(e)


async def leave_all_chats_from_account(
    client: TelegramClient, phone: str, on_progress=None, cancel_event=None
) -> dict:
    left = 0
    failed = 0
    errors = []
    try:
        async for d in client.iter_dialogs():
            if cancel_event and cancel_event.is_set():
                break
            if isinstance(d.entity, (Channel, Chat)):
                ok, msg = await _leave_chat(client, d.entity)
                chat_name = getattr(d, "name", "") or str(d.id)
                if ok:
                    left += 1
                else:
                    failed += 1
                    errors.append((chat_name, msg))
                if on_progress:
                    on_progress(left, failed)
                await asyncio.sleep(1)
            elif isinstance(d.entity, User) and d.is_user and d.entity.is_self:
                pass
    except Exception as e:
        return {"phone": phone, "left": left, "failed": failed, "errors": errors + [("", str(e))]}
    return {"phone": phone, "left": left, "failed": failed, "errors": errors}


async def run_leave_all_chats(on_progress=None, cancel_event=None) -> list[dict]:
    cfg = load_config()
    if not cfg or not cfg.get("api_id") or not cfg.get("api_hash"):
        raise ValueError("Не настроены api_id и api_hash")
    accounts = [a for a in cfg["accounts"] if is_account_authorized(a["phone"])]
    if not accounts:
        raise ValueError("Нет авторизованных аккаунтов")
    account_stats = {}

    async def leave_for_account(acc: dict):
        phone = acc["phone"]
        session_path = SESSIONS_DIR / phone.replace("+", "").replace(" ", "")
        proxy = acc.get("proxy") or cfg.get("proxy")
        client = create_client(session_path, cfg["api_id"], cfg["api_hash"], proxy)
        account_stats[phone] = {"left": 0, "failed": 0}
        try:
            await client.connect()
            if not await client.is_user_authorized():
                return {"phone": phone, "left": 0, "failed": 0, "errors": [("", "Не авторизован")]}

            def acc_progress(left, failed):
                account_stats[phone]["left"] = left
                account_stats[phone]["failed"] = failed
                if on_progress:
                    total_l = sum(s["left"] for s in account_stats.values())
                    total_f = sum(s["failed"] for s in account_stats.values())
                    on_progress(total_l, total_f)

            r = await leave_all_chats_from_account(client, phone, on_progress=acc_progress, cancel_event=cancel_event)
            return {"phone": phone, **r}
        except Exception as e:
            return {"phone": phone, "left": 0, "failed": 0, "errors": [("", str(e))]}
        finally:
            await client.disconnect()

    results = await asyncio.gather(*[leave_for_account(acc) for acc in accounts])
    return list(results)
