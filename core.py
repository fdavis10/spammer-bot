import json
import os
import shutil
import sys
import uuid
from pathlib import Path

from openpyxl import Workbook

def _get_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent

BASE_DIR = _get_base_dir()
CONFIG_PATH = BASE_DIR / "config.json"
DATA_DIR = BASE_DIR / "data"
ATTACHMENTS_DIR = DATA_DIR / "attachments"
TDATA_IMPORT_DIR = DATA_DIR / "tdata_import"


def add_file_to_storage(source_path: str) -> str:
    src = Path(source_path)
    if not src.exists():
        raise FileNotFoundError(f"Файл не найден: {source_path}")
    ATTACHMENTS_DIR.mkdir(parents=True, exist_ok=True)
    ext = src.suffix or ""
    name = f"{uuid.uuid4().hex}{ext}"
    dest = ATTACHMENTS_DIR / name
    shutil.copy2(src, dest)
    return str(dest.resolve())

_base = Path(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")))
SESSIONS_DIR = _base / "spammer-bot" / "sessions"
MAX_ACCOUNTS = 10

OLD_SESSIONS = BASE_DIR / "sessions"


def _migrate_sessions():
    if not OLD_SESSIONS.exists():
        return
    SESSIONS_DIR.mkdir(parents=True, exist_ok=True)
    for f in OLD_SESSIONS.glob("*.session*"):
        dst = SESSIONS_DIR / f.name
        if not dst.exists() or f.stat().st_mtime > dst.stat().st_mtime:
            import shutil

            shutil.copy2(f, dst)


def load_config():
    if not CONFIG_PATH.exists():
        return None
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


def save_config(cfg):
    DATA_DIR.mkdir(exist_ok=True)
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=4)


def is_account_authorized(phone):
    name = phone.replace("+", "").replace(" ", "")
    return (SESSIONS_DIR / f"{name}.session").exists()


def get_templates():
    cfg = load_config() or {}
    return cfg.get("templates") or []


def save_templates(templates):
    cfg = load_config() or {}
    cfg["templates"] = templates
    save_config(cfg)


def add_template(name, message, parse_mode, attachments):
    tpl = {"name": name.strip(), "message": message or "", "parse_mode": parse_mode or "none", "attachments": list(attachments or [])}
    templates = get_templates()
    templates.append(tpl)
    save_templates(templates)


def delete_template(name):
    templates = [t for t in get_templates() if t.get("name") != name]
    save_templates(templates)


def set_account_premium(phone: str, premium: bool) -> None:
    c = load_config() or {}
    for acc in c.get("accounts", []):
        if acc.get("phone") == phone:
            acc["premium"] = bool(premium)
            break
    save_config(c)


CHAT_LINKS_PATH = DATA_DIR / "chat_links.json"
JOINED_CHATS_PATH = DATA_DIR / "joined_chats.json"
SENT_MESSAGE_LINKS_PATH = DATA_DIR / "sent_message_links.json"


def _normalize_phone(phone: str) -> str:
    return str(phone or "").replace("+", "").replace(" ", "").strip()


def _normalize_link(link: str) -> str:
    s = str(link or "").strip()
    for prefix in ("https://", "http://"):
        if s.lower().startswith(prefix):
            s = s[len(prefix):]
    if s.lower().startswith("telegram.me/"):
        s = "t.me/" + s[12:]
    elif s.lower().startswith("telegram.dog/"):
        s = "t.me/" + s[13:]
    return s.strip("/") or ""


def _load_joined_chats() -> dict:
    if not JOINED_CHATS_PATH.exists():
        return {}
    try:
        with open(JOINED_CHATS_PATH, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, IOError):
        return {}


def _save_joined_chats(data: dict) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    with open(JOINED_CHATS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=0)


def get_joined_links(phone: str) -> set[str]:
    key = _normalize_phone(phone)
    data = _load_joined_chats()
    links = data.get(key) or []
    return set(_normalize_link(l) for l in links if l)


def add_joined_links(phone: str, links: list[str]) -> None:
    if not links:
        return
    key = _normalize_phone(phone)
    data = _load_joined_chats()
    current = set(data.get(key) or [])
    for link in links:
        nl = _normalize_link(link)
        if nl:
            current.add(nl)
    data[key] = list(current)
    _save_joined_chats(data)


def clear_joined_for_account(phone: str) -> None:
    key = _normalize_phone(phone)
    data = _load_joined_chats()
    if key in data:
        del data[key]
        _save_joined_chats(data)


def export_joined_chats_to_xlsx(filepath: str) -> None:
    cfg = load_config() or {}
    accounts = cfg.get("accounts") or []
    if not accounts:
        raise ValueError("Нет аккаунтов в настройках")

    columns_data = []
    for acc in accounts:
        phone = str(acc.get("phone") or "").strip()
        links = sorted(get_joined_links(phone))
        columns_data.append((phone or "(без номера)", list(links)))

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Чаты")
    else:
        ws.title = "Чаты"

    for col_idx, (phone, links) in enumerate(columns_data, start=1):
        ws.cell(row=1, column=col_idx, value=phone)
        for row_idx, link in enumerate(links, start=2):
            ws.cell(row=row_idx, column=col_idx, value=link)

    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)


def _load_sent_message_links() -> dict:
    if not SENT_MESSAGE_LINKS_PATH.exists():
        return {}
    try:
        with open(SENT_MESSAGE_LINKS_PATH, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, IOError):
        return {}


def _save_sent_message_links(data: dict) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    with open(SENT_MESSAGE_LINKS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=0)


def add_sent_message_link(phone: str, link: str) -> None:
    if not link or not str(link).strip():
        return
    link = str(link).strip()
    key = _normalize_phone(phone)
    data = _load_sent_message_links()
    current = data.get(key) or []
    if link not in current:
        current.append(link)
        data[key] = current
        _save_sent_message_links(data)


def get_sent_message_links(phone: str) -> list[str]:
    key = _normalize_phone(phone)
    data = _load_sent_message_links()
    return list(data.get(key) or [])


def export_sent_message_links_to_xlsx(filepath: str) -> None:
    from openpyxl.styles import Font

    cfg = load_config() or {}
    accounts = cfg.get("accounts") or []
    if not accounts:
        raise ValueError("Нет аккаунтов в настройках")

    columns_data = []
    for acc in accounts:
        phone = str(acc.get("phone") or "").strip()
        links = get_sent_message_links(phone)
        columns_data.append((phone or "(без номера)", links))

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Ссылки на сообщения")
    else:
        ws.title = "Ссылки на сообщения"

    link_font = Font(color="0563C1", underline="single")
    col_width = 45

    for col_idx, (phone, links) in enumerate(columns_data, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = col_width
        ws.cell(row=1, column=col_idx, value=phone)
        for row_idx, link in enumerate(links, start=2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = "Ссылка" if link else ""
            if link:
                full_url = link if link.startswith("http") else f"https://{link}"
                cell.hyperlink = full_url
                cell.font = link_font

    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)


def _migrate_chat_links():
    cfg = load_config()
    if not cfg or "chat_links" not in cfg:
        return
    links = cfg.get("chat_links") or []
    if links:
        DATA_DIR.mkdir(exist_ok=True)
        with open(CHAT_LINKS_PATH, "w", encoding="utf-8") as f:
            json.dump(links, f, ensure_ascii=False, indent=0)
    del cfg["chat_links"]
    save_config(cfg)


def get_chat_links():
    if not CHAT_LINKS_PATH.exists():
        return []
    try:
        with open(CHAT_LINKS_PATH, encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return []


def save_chat_links(links: list[str]):
    DATA_DIR.mkdir(exist_ok=True)
    cleaned = list(dict.fromkeys(str(l).strip() for l in links if l and str(l).strip()))
    with open(CHAT_LINKS_PATH, "w", encoding="utf-8") as f:
        json.dump(cleaned, f, ensure_ascii=False, indent=0)


def add_chat_links(new_links: list[str]):
    current = get_chat_links()
    seen = set(current)
    for link in new_links:
        l = str(link).strip()
        if l and l not in seen:
            seen.add(l)
            current.append(l)
    save_chat_links(current)


def _simplify_log_line(line: str) -> str | None:
    s = line.strip()
    if "ОТПРАВЛЕНО" in s:
        if " -> " in s:
            chat = s.split(" -> ", 1)[-1].split(":")[0].strip()
            return f"Отправлено: {chat}"
        return "Отправлено"
    if "ОШИБКА" in s:
        if " -> " in s:
            rest = s.split(" -> ", 1)[-1]
            chat = rest.split(":")[0].strip() if ":" in rest else rest
            return f"Ошибка: {chat}"
        return "Ошибка"
    if "Готово" in s and "отправлено" in s:
        return s.split("|")[-1].strip()
    return None


def get_join_stats() -> int:
    data = _load_joined_chats() or {}
    all_links = set()
    for links in data.values():
        for l in links or []:
            nl = _normalize_link(l)
            if nl:
                all_links.add(nl)
    return len(all_links)


def add_join_stats(count: int) -> None:
    return None


BROADCAST_LOG_PATH = DATA_DIR / "broadcast.log"


def clear_broadcast_log() -> None:
    if BROADCAST_LOG_PATH.exists():
        BROADCAST_LOG_PATH.write_text("", encoding="utf-8")


def get_stats():
    log_path = BROADCAST_LOG_PATH
    if not log_path.exists():
        return {"total_sent": 0, "total_failed": 0, "last_lines": []}
    sent = 0
    failed = 0
    simplified = []
    with open(log_path, encoding="utf-8") as f:
        for line in f:
            raw = line.strip()
            if "ОТПРАВЛЕНО" in raw:
                sent += 1
            elif "ОШИБКА" in raw:
                failed += 1
            short = _simplify_log_line(raw)
            if short:
                simplified.append(short)
    return {
        "total_sent": sent,
        "total_failed": failed,
        "last_lines": simplified[-50:] if len(simplified) > 50 else simplified,
    }


_migrate_sessions()
_migrate_chat_links()
